import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy
import os

Fecha = str(input("Colocar la fecha de Inicio (dd/mm/aaaa): "))

def Semana():
    Personal = ['Bernardo Conquet', 'Miguel Palmera', 'Jairo Peña', 'Henry Flores', 'Jairo Prieto', 'Juan Gil', 'Jackson Diaz', 'Anderson Ramirez', 'Miguel Contreras']
    Mañana = ['Bernardo Conquet', 'Jairo Peña', 'Henry Flores', 'Jairo Prieto', 'Anderson Ramirez']
    Tarde = ['Bernardo Conquet', 'Miguel Palmera', 'Jairo Prieto', 'Juan Gil', 'Miguel Contreras', 'Jackson Diaz']
    Vacaciones = ['Anderson Ramirez', 'Jackson Diaz']

    Semana_Mañana = []
    Semana_Tarde = []

    for i in range(5):
        m = True
        t = True
        while m:
            M = random.randint(0, len(Personal)-1)
            if (Personal[M] in Mañana) and (Personal[M] not in Semana_Mañana) and (Personal[M] not in Vacaciones):
                Semana_Mañana.append(Personal[M])
                m = False
            elif (Personal[M] in Mañana) and (len(Mañana) - len(Vacaciones) < 5) and (Personal[M] not in Vacaciones) and len(Semana_Mañana) == 4 and (Semana_Mañana[i - 1] != Personal[M]):
                Semana_Mañana.append(Personal[M])
                m = False
            else:
                m = True

        while t:
            T = random.randint(0, len(Personal)-1)
            if (Personal[T] in Tarde) and (Personal[T] not in Semana_Tarde) and (Personal[T] not in Vacaciones) and (M != T):
                Semana_Tarde.append(Personal[T])
                t = False
            elif (Personal[T] in Tarde) and (len(Tarde) - len(Vacaciones) < 5) and (Personal[T] not in Vacaciones) and len(Semana_Tarde) == 4 and (Semana_Tarde[i - 1] != Personal[T]):
                Semana_Tarde.append(Personal[T])
                t = False
            else:
                t = True

    print(Semana_Mañana)
    print(Semana_Tarde)
    print('______________________________________________________________________________________')
    return  Semana_Mañana, Semana_Tarde

SM1, ST1 = Semana()
SM2, ST2 = Semana()
SM3, ST3 = Semana()
SM4, ST4 = Semana()

wb = Workbook()
ws = wb.active
ft1 = Font(name = 'Arial', size = 28, bold = True)
ft2 = copy(ft1)
ft2.size = "18"

ws['A1'] = 'Cronograma de Entrega y Recepcion de Llaves al Vigilante CNTO'
ws.merge_cells('A1:H1')
ws['A1'].font = ft1
ws['A1'].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
ws.row_dimensions[1].height = 70

Semana = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']
ws.append(Semana)
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 20

ws.insert_rows(2)
ws.move_range("A3:E3", cols=3)

#Semana 1
Recibe = [Fecha, 'al', 'Recibe', SM1[0], SM1[1], SM1[2], SM1[3], SM1[4]]
ws.append(Recibe)
Entrega = ['=A4+4', '', 'Entrega', ST1[0], ST1[1], ST1[2], ST1[3], ST1[4]]
ws.append(Entrega)
#Semana 2
Recibe = ['=A5+3', 'al', 'Recibe', SM2[0], SM2[1], SM2[2], SM2[3], SM2[4]]
ws.append(Recibe)
Entrega = ['=A7+4', '', 'Entrega', ST2[0], ST2[1], ST2[2], ST2[3], ST2[4]]
ws.append(Entrega)
#Semana 3
Recibe = ['=A8+3', 'al', 'Recibe', SM3[0], SM3[1], SM3[2], SM3[3], SM3[4]]
ws.append(Recibe)
Entrega = ['=A10+4', '', 'Entrega', ST3[0], ST3[1], ST3[2], ST3[3], ST3[4]]
ws.append(Entrega)
#Semana 4
Recibe = ['=A11+3', 'al', 'Recibe', SM4[0], SM4[1], SM4[2], SM4[3], SM4[4]]
ws.append(Recibe)
Entrega = ['=A13+4', '', 'Entrega', ST4[0], ST4[1], ST4[2], ST4[3], ST4[4]]
ws.append(Entrega)

ws.move_range('A10:H11', rows = 3)
ws.move_range('A8:H9', rows = 2)
ws.move_range('A6:H7', rows = 1)

ws.merge_cells('B4:B5')
ws.merge_cells('B7:B8')
ws.merge_cells('B10:B11')
ws.merge_cells('B13:B14')

ws['B4'].alignment = Alignment(horizontal = "center", vertical = "center")
ws['B7'].alignment = Alignment(horizontal = "center", vertical = "center")
ws['B10'].alignment = Alignment(horizontal = "center", vertical = "center")
ws['B13'].alignment = Alignment(horizontal = "center", vertical = "center")

ws.row_dimensions[4].height = 47.25
ws.row_dimensions[5].height = 47.25
ws.row_dimensions[7].height = 47.25
ws.row_dimensions[8].height = 47.25
ws.row_dimensions[10].height = 47.25
ws.row_dimensions[11].height = 47.25
ws.row_dimensions[13].height = 47.25
ws.row_dimensions[14].height = 47.25

for x in range(2, 15):
    for y in range(1, 9):
        ws.cell(row = x, column = y).font = ft2
        ws.cell(row = x, column = y).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)

for x in range(4, 15):
    for y in range(4, 9):
        ws.cell(row = x, column = y).font = Font(name = 'Arial', size = 18, bold = False)

R = [4, 7, 10, 13]
for x in R:
    for y in range(3, 9):
        ws.cell(row = x, column = y).fill = PatternFill("solid", fgColor = "BFBFBF")

thin = Side(border_style = "thin", color = "000000")

for y in range(4, 9):
    ws.cell(row = 3, column = y).border = Border(top = thin, left= thin, right = thin, bottom = thin)

RE = [4, 5, 7, 8, 10, 11, 13, 14]
for x in RE:
    for y in range(1, 9):
        ws.cell(row = x, column = y).border = Border(top = thin, left = thin, right = thin, bottom = thin)

ws.column_dimensions['A'].width = 19
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 19
ws.column_dimensions['E'].width = 19
ws.column_dimensions['F'].width = 19
ws.column_dimensions['G'].width = 19
ws.column_dimensions['H'].width = 19

# Save the file
wb.save("Python Scripts/Cronograma_Vigilancia.xlsx")

os.system('start "excel" "Python Scripts/Cronograma_Vigilancia.xlsx"')
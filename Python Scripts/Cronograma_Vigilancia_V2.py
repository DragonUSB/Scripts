import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from copy import copy

Fecha = str(input("Colocar la fecha de Inicio (dd/mm/aaaa): "))

Personal = ['Bernardo Conquet', 'Miguel Palmera', 'Jairo Peña', 'Henry Flores', 'Jairo Prieto', 'Juan Gil', 'Jackson Diaz', 'Anderson Ramirez', 'Miguel Contreras']
Mañana = ['Bernardo Conquet', 'Jairo Peña', 'Henry Flores', 'Jairo Prieto', 'Anderson Ramirez']
Tarde = ['Bernardo Conquet', 'Miguel Palmera', 'Jairo Prieto', 'Juan Gil', 'Miguel Contreras', 'Jackson Diaz']
Vacaciones = ['Anderson Ramirez', 'Jackson Diaz']

Semana_Mañana = []
Semana_Tarde = []

for i in range(5):
    m = True
    t = True
    while m:
        M = random.randint(0, len(Personal)-1)
        if (Personal[M] in Mañana) and (Personal[M] not in Semana_Mañana) and (Personal[M] not in Vacaciones):
            Semana_Mañana.append(Personal[M])
            m = False
        elif (Personal[M] in Mañana) and (len(Mañana) - len(Vacaciones) < 5) and (Personal[M] not in Vacaciones) and len(Semana_Mañana) == 4 and (Semana_Mañana[i - 1] != Personal[M]):
            Semana_Mañana.append(Personal[M])
            m = False
        else:
            m = True

    while t:
        T = random.randint(0, len(Personal)-1)
        if (Personal[T] in Tarde) and (Personal[T] not in Semana_Tarde) and (Personal[T] not in Vacaciones) and (M != T):
            Semana_Tarde.append(Personal[T])
            t = False
        elif (Personal[T] in Tarde) and (len(Tarde) - len(Vacaciones) < 5) and (Personal[T] not in Vacaciones) and len(Semana_Tarde) == 4 and (Semana_Tarde[i - 1] != Personal[T]):
            Semana_Tarde.append(Personal[T])
            t = False
        else:
            t = True

print(Semana_Mañana)
print(Semana_Tarde)

wb = Workbook()
ws = wb.active
ft1 = Font(name = 'Arial', size = 28, bold = True)
ft2 = copy(ft1)
ft2.size = "18"

ws['A1'] = 'Cronograma de Entrega y Recepcion de Llaves al Vigilante CNTO'
ws.merge_cells('A1:H1')
ws['A1'].font = ft1
ws['A1'].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
ws.row_dimensions[1].height = 70

Semana = ['Lunes', 'Martes', 'Miercoles', 'Jueves', 'Viernes']
ws.append(Semana)
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 20

ws.insert_rows(2)
ws.move_range("A3:E3", cols=3)

Recibe = [Fecha, 'al', 'Recibe', Semana_Mañana[0], Semana_Mañana[1], Semana_Mañana[2], Semana_Mañana[3], Semana_Mañana[4]]
ws.append(Recibe)
Entrega = ['=A4+4', '', 'Entrega', Semana_Tarde[0], Semana_Tarde[1], Semana_Tarde[2], Semana_Tarde[3], Semana_Tarde[4]]
ws.append(Entrega)
ws.merge_cells('B4:B5')
ws['B4'].alignment = Alignment(horizontal = "center", vertical = "center")
ws.row_dimensions[4].height = 47.25
ws.row_dimensions[5].height = 47.25

for x in range(2, 15):
    for y in range(1, 9):
        ws.cell(row = x, column = y).font = ft2
        ws.cell(row = x, column = y).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)

for x in range(4, 6):
    for y in range(4, 9):
        ws.cell(row = x, column = y).font = Font(name = 'Arial', size = 18, bold = False)

ws.column_dimensions['A'].width = 19
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 19
ws.column_dimensions['E'].width = 19
ws.column_dimensions['F'].width = 19
ws.column_dimensions['G'].width = 19
ws.column_dimensions['H'].width = 19

# Save the file
wb.save("Python Scripts/Cronograma_Vigilancia.xlsx")